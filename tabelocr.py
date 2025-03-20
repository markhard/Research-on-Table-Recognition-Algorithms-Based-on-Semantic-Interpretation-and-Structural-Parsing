from paddleocr import PaddleOCR, draw_ocr
from PIL import Image
import time
import numpy as np
from openpyxl import load_workbook

def find_row(data, tolerance, min_count):
    """Find the minimum value of each contiguous region in the data, allowing for fault-tolerant intervals and a number of occurrences greater than 4 in each interval.“”"”
        :param data: list of input data
        :param tolerance: fault tolerance, maximum tolerated interval difference
        :param min_count: the number of occurrences of a number in each consecutive region must be greater than this value to be considered a valid region
        :return: local minimum value for each contiguous region
        """
    data = sorted(data)
    # print(data)
    minima = []
    start_idx = 0
    for i in range(1, len(data)):

        if data[i] - data[i - 1] > tolerance:
            if (i - start_idx) >= min_count:
                minima.append(data[start_idx])
            start_idx = i
    if (len(data) - start_idx) >= min_count:
        minima.append(data[start_idx])
    minima = [round(num) for num in minima]
    return minima

start_time = time.time()
ocr = PaddleOCR(use_angle_cls=True, lang="ch")  # need to run only once to download and load model into memory
ex_image_path = 'img/demo2.jpg'

tabelhead=0
input_image = Image.open(ex_image_path)
results = ocr.ocr(ex_image_path, cls=False)[0]
print(results)
leftnum = [results[i][0][0][0] for i in range(len(results))]
print(leftnum)
rightnum = [results[i][0][1][0] for i in range(len(results))]
print(sorted(rightnum))
topnum = [results[i][0][0][1] for i in range(len(results))]
tolerance=results[0][0][2][1]-results[0][0][0][1]
for i in range(len(results)):
    if abs(results[0][0][0][1]-results[i][0][0][1])<tolerance:
        tabelhead=tabelhead+1
    else:
        break
print(tabelhead)
a=find_row(leftnum,tolerance=tolerance, min_count=tabelhead-1)
print(len(a),a)

vertical=[0]*(tabelhead-1)
# top=[0]*tabelhead
for i in range(tabelhead-1):
    filtered_numbers_left = [num for num in rightnum if rightnum[i] <= num <= leftnum[i + 1]]
    filtered_numbers_right = [num for num in leftnum if rightnum[i] <= num <= leftnum[i + 1]]
    vertical[i]=round((min(filtered_numbers_right)+max(filtered_numbers_left))/2)
print(vertical)
left =[0] + vertical#Left
right= vertical + [input_image.width]#right
row = find_row(topnum,tolerance=tolerance/2, min_count=tabelhead/2)
row=row+[input_image.height]
print(row)
#Create a blank table
table_dates = [[] for _ in range(tabelhead)]
for e in table_dates:
    for _ in range(len(row)-1):
        e.append('')
found=False
#multi-line text issues
for i in range(len(results)):
    for j in range(tabelhead):
        for k in range(len(row)-1):
            if results[i][0][0][0]>=left[j] and results[i][0][1][0]<=right[j] and results[i][0][0][1] >= row[k] and results[i][0][0][1]<row[k+1]:
                if table_dates[j][k] is None:
                    table_dates[j][k] = ""
                table_dates[j][k] = table_dates[j][k]+results[i][1][0]
                found=True
                break
            if j<tabelhead-1 and k <len(row)-1:
                if (results[i][0][0][0]>left[j] and results[i][0][1][0]<right[j+1] and results[i][0][0][1] >= row[k] and results[i][0][0][1]<row[k+1]
                      and results[i][0][0][0]<right[j] and results[i][0][1][0]>left[j+1]):
                    table_dates[j][k] = table_dates[j][k] + results[i][1][0]
                    table_dates[j+1][k] = None
                    found = True
                    break
        if found==True:
            found=False
            break

for date in table_dates:
    print(date)
workbook = load_workbook(filename="img/demo.xlsx")
sheet = workbook.active
for j in range(tabelhead):
    for k in range(len(row)-1):
        if table_dates[j][k]==None:
            sheet.merge_cells(
                start_row=k+1, start_column=j,
                end_row=k+1, end_column=j+1
            )
        else:
            sheet.cell(row=k+1, column=j+1).value = table_dates[j][k]
workbook.save("img/demo.xlsx")
end_time = time.time()

elapsed_time = end_time - start_time
print(f"Code Runtime：{elapsed_time}s")
